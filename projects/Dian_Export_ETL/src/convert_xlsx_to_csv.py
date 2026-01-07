from __future__ import annotations
from pathlib import Path
from typing import List
import csv
from openpyxl import load_workbook


def convertir_xlsx_nuevos_a_csv(
    *,
    extracted_dir: str | Path,
    csv_dir: str | Path,
    log_file: str | Path | None = None,
    log_dir: str | Path | None = None,
    verbose: bool = True,
) -> List[Path]:

    extracted_dir = Path(extracted_dir)
    csv_dir = Path(csv_dir)
    csv_dir.mkdir(parents=True, exist_ok=True)

    # Resolver log_file desde log_dir
    if log_file is None:
        if log_dir is None:
            # usa un directorio "logs" a nivel de repo o relativo
            log_dir = (csv_dir.parent.parent / "logs") if (csv_dir.parent.parent).exists() else (csv_dir.parent / "logs")
        log_dir = Path(log_dir)
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file = log_dir / "xlsx_to_csv.log"
    else:
        log_file = Path(log_file)
        log_file.parent.mkdir(parents=True, exist_ok=True)

    # --- Cargar log de archivos ya convertidos ---
    if log_file.exists():
        procesados = {
            line.strip()
            for line in log_file.read_text(encoding="utf-8").splitlines()
            if line.strip()
        }
    else:
        procesados = set()

    if verbose:
        print(f"Log de control: {log_file}")
        print(f"Archivos ya convertidos: {len(procesados)}")

    # --- Buscar TODOS los .xlsx en extracted (recursivo) ---
    xlsx_files = sorted(extracted_dir.rglob("*.xlsx"))
    if verbose:
        print(f"XLSX encontrados en disco: {len(xlsx_files)}")

    # Filtrar solo nuevos
    nuevos = [p for p in xlsx_files if str(p.resolve()) not in procesados]

    if verbose:
        print(f"XLSX nuevos a procesar: {len(nuevos)}")

    csv_generados: List[Path] = []

    for xlsx_path in nuevos:
        if verbose:
            print(f"\nConvirtiendo {xlsx_path.name}...")

        try:
            wb = load_workbook(
                filename=xlsx_path,
                read_only=True,
                data_only=True
            )
            ws = wb.active

            out_csv = csv_dir / f"{xlsx_path.stem}.csv"

            with out_csv.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)

            wb.close()

            if verbose:
                print(f"   CSV guardado en: {out_csv}")

            csv_generados.append(out_csv)
            procesados.add(str(xlsx_path.resolve()))

        except Exception as e:
            if verbose:
                print(f"   Error convirtiendo {xlsx_path}: {e}")

    # --- Guardar log actualizado ---
    log_file.write_text("\n".join(sorted(procesados)), encoding="utf-8")

    if verbose:
        print(f"\nCSV generados en esta corrida: {len(csv_generados)}")

    return csv_generados


